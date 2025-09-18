# openpyxl模块是用于处理Excel文件的第三方库
# 可以对Excel文件中的数据进行写入和读取

from module_get_attraction import * # 里面的requests和re也会被import
import openpyxl

html = get_html('https://www.weather.com.cn/weather1d/101020100.shtml')
lst = parse_html(html)

# 创建一个新的excel工作簿
workbook = openpyxl.Workbook()
# 在excel文件中创建工作表
sheet = workbook.create_sheet('Weather') # sheet名
# 向工作表中添加数据
for item in lst:
    sheet.append(item)
# 保存工作簿
workbook.save('Weather of Attractions.xlsx') # 文件名

# 从excel中读取数据
wb = openpyxl.load_workbook('Weather of Attractions.xlsx')
sheet = workbook['Weather']
# 表格数据是二维列表, 先遍历的是行, 后遍历列
lst = [] # 存储行数据
for row in sheet.rows:
    sublst = [] # 存储单元格数据
    for cell in row:
        sublst.append(cell.value)
    lst.append(sublst)

print(lst)